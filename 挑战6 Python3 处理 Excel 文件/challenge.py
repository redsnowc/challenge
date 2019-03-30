import openpyxl
import datetime

filename = 'courses.xlsx'
wb = openpyxl.load_workbook(filename)


def combine_excel():
    sheet_1 = wb[wb.sheetnames[0]]
    sheet_2 = wb[wb.sheetnames[1]]
    sheet_3 = wb.create_sheet('combine')
    for value_1 in sheet_1.values:
        for value_2 in sheet_2.values:
            if value_1[1] == value_2[1]:
                sheet_3.append(list(value_1) + [value_2[2]])
    wb.save(filename)


def split_excel():
    sheet = wb[wb.sheetnames[2]]
    time = []
    first_row = []
    for row_cell in sheet['A1':'D1']:
        for cell in row_cell:
            first_row.append(cell.value)

    for value in sheet.values:
        if type(value[0]) != type('test'):
            time.append(value[0].strftime('%Y'))

    for name in set(time):
        split_excel = openpyxl.Workbook()
        split_excel_sheet = split_excel.active
        split_excel_sheet.title = name
        split_excel_sheet.append(first_row)

        for time_value in sheet.values:
            if type(time_value[0]) != type('test'):
                if time_value[0].strftime('%Y') == name:
                    split_excel_sheet.append(time_value)

        split_excel.save('%s.xlsx' % (name))


combine_excel()
split_excel()
